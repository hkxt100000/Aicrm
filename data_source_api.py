"""
源数据管理 API
支持灵活的数据源配置和数据推送
"""
import uuid
import secrets
import json
import time
from typing import List, Dict, Optional, Any
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, HTTPException, Header, File, UploadFile, Depends
from pydantic import BaseModel
import sqlite3

# Excel 处理
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/data-source", tags=["源数据管理"])

# ========== 数据模型 ==========

class DataSourceCreate(BaseModel):
    """创建数据源"""
    name: str                          # 数据源名称
    source_type: str                   # 数据源类型：order/product/supplier/custom
    description: Optional[str] = None  # 描述
    field_schema: Dict                 # 字段定义
    auto_sync: bool = False            # 是否自动同步
    sync_interval: int = 3600          # 同步间隔（秒）

class DataSourceUpdate(BaseModel):
    """更新数据源"""
    name: Optional[str] = None
    description: Optional[str] = None
    field_schema: Optional[Dict] = None
    auto_sync: Optional[bool] = None
    sync_interval: Optional[int] = None
    status: Optional[str] = None

class DataPushRequest(BaseModel):
    """数据推送请求"""
    data: List[Dict]                   # 数据列表
    incremental: bool = False          # 是否增量更新
    data_key_field: Optional[str] = None  # 数据主键字段名

# ========== 数据库操作 ==========

def get_db():
    """获取数据库连接"""
    from config import DB_PATH
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_data_source_tables():
    """初始化数据源管理表"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 读取并执行建表SQL
    with open('schema_data_source.sql', 'r', encoding='utf-8') as f:
        sql_script = f.read()
        cursor.executescript(sql_script)
    
    conn.commit()
    conn.close()

def generate_api_key() -> str:
    """生成API密钥"""
    return f"thc_{secrets.token_urlsafe(32)}"

# ========== API 端点 ==========

@router.post("/create")
def create_data_source(data: DataSourceCreate):
    """创建数据源"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 生成ID和API密钥
        source_id = f"ds_{uuid.uuid4().hex[:12]}"
        api_key = generate_api_key()
        now = int(time.time() * 1000)
        
        # 插入数据源
        cursor.execute("""
            INSERT INTO data_sources (
                id, name, source_type, description, api_key, status,
                auto_sync, sync_interval, field_schema,
                total_records, sync_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            source_id, data.name, data.source_type, data.description,
            api_key, 'active', data.auto_sync, data.sync_interval,
            json.dumps(data.field_schema, ensure_ascii=False),
            0, 0, now, now
        ))
        
        conn.commit()
        
        # 返回创建的数据源信息
        result = {
            "id": source_id,
            "name": data.name,
            "source_type": data.source_type,
            "api_key": api_key,
            "status": "active",
            "push_url": f"/api/data-source/push",
            "created_at": now
        }
        
        conn.close()
        return {"code": 0, "message": "数据源创建成功", "data": result}
        
    except Exception as e:
        return {"code": 1, "message": f"创建失败：{str(e)}"}

@router.get("/list")
def list_data_sources(
    source_type: Optional[str] = None,
    status: Optional[str] = None
):
    """获取数据源列表"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 构建查询
        query = "SELECT * FROM data_sources WHERE deleted = 0"
        params = []
        
        if source_type:
            query += " AND source_type = ?"
            params.append(source_type)
        
        if status:
            query += " AND status = ?"
            params.append(status)
        
        query += " ORDER BY created_at DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # 转换为字典列表
        result = []
        for row in rows:
            item = dict(row)
            # 解析 JSON 字段
            if item.get('field_schema'):
                item['field_schema'] = json.loads(item['field_schema'])
            result.append(item)
        
        conn.close()
        return {"code": 0, "data": result, "total": len(result)}
        
    except Exception as e:
        return {"code": 1, "message": f"查询失败：{str(e)}"}

@router.get("/{source_id}")
def get_data_source(source_id: str):
    """获取数据源详情"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM data_sources WHERE id = ? AND deleted = 0
        """, (source_id,))
        
        row = cursor.fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="数据源不存在")
        
        result = dict(row)
        if result.get('field_schema'):
            result['field_schema'] = json.loads(result['field_schema'])
        
        conn.close()
        return {"code": 0, "data": result}
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"查询失败：{str(e)}"}

@router.put("/{source_id}")
def update_data_source(source_id: str, data: DataSourceUpdate):
    """更新数据源"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 检查数据源是否存在
        cursor.execute("""
            SELECT id FROM data_sources WHERE id = ? AND deleted = 0
        """, (source_id,))
        
        if not cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="数据源不存在")
        
        # 构建更新语句
        updates = []
        params = []
        
        if data.name is not None:
            updates.append("name = ?")
            params.append(data.name)
        
        if data.description is not None:
            updates.append("description = ?")
            params.append(data.description)
        
        if data.field_schema is not None:
            updates.append("field_schema = ?")
            params.append(json.dumps(data.field_schema, ensure_ascii=False))
        
        if data.auto_sync is not None:
            updates.append("auto_sync = ?")
            params.append(data.auto_sync)
        
        if data.sync_interval is not None:
            updates.append("sync_interval = ?")
            params.append(data.sync_interval)
        
        if data.status is not None:
            updates.append("status = ?")
            params.append(data.status)
        
        if not updates:
            conn.close()
            return {"code": 0, "message": "无更新内容"}
        
        # 添加更新时间
        updates.append("updated_at = ?")
        params.append(int(time.time() * 1000))
        params.append(source_id)
        
        # 执行更新
        cursor.execute(f"""
            UPDATE data_sources SET {', '.join(updates)}
            WHERE id = ?
        """, params)
        
        conn.commit()
        conn.close()
        
        return {"code": 0, "message": "更新成功"}
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"更新失败：{str(e)}"}

@router.delete("/{source_id}")
def delete_data_source(source_id: str):
    """删除数据源（软删除）"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE data_sources SET deleted = 1, updated_at = ?
            WHERE id = ? AND deleted = 0
        """, (int(time.time() * 1000), source_id))
        
        if cursor.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="数据源不存在")
        
        conn.commit()
        conn.close()
        
        return {"code": 0, "message": "删除成功"}
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"删除失败：{str(e)}"}

@router.post("/{source_id}/regenerate-key")
def regenerate_api_key(source_id: str):
    """重新生成API密钥"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 生成新密钥
        new_api_key = generate_api_key()
        
        cursor.execute("""
            UPDATE data_sources SET api_key = ?, updated_at = ?
            WHERE id = ? AND deleted = 0
        """, (new_api_key, int(time.time() * 1000), source_id))
        
        if cursor.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="数据源不存在")
        
        conn.commit()
        conn.close()
        
        return {"code": 0, "message": "密钥已重新生成", "data": {"api_key": new_api_key}}
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"操作失败：{str(e)}"}

class BatchClearRequest(BaseModel):
    """批量清理请求"""
    clear_type: str  # 'by_time' 或 'all'
    days: Optional[int] = None  # 当 clear_type='by_time' 时需要

@router.post("/{source_id}/batch-clear")
def batch_clear_records(source_id: str, request: BatchClearRequest):
    """批量清理数据记录
    
    Args:
        source_id: 数据源ID
        request: 清理请求
            - clear_type: 清理类型 ('by_time' 按时间清空, 'all' 全部清空)
            - days: 保留最近N天的数据（clear_type='by_time'时需要）
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 验证数据源是否存在
        cursor.execute("""
            SELECT id, name, total_records FROM data_sources
            WHERE id = ? AND deleted = 0
        """, (source_id,))
        
        source_row = cursor.fetchone()
        if not source_row:
            conn.close()
            raise HTTPException(status_code=404, detail="数据源不存在")
        
        source = dict(source_row)
        deleted_count = 0
        
        if request.clear_type == 'by_time':
            # 按时间清空：删除指定天数之前的记录
            if not request.days:
                conn.close()
                raise HTTPException(status_code=400, detail="缺少参数：days")
            
            # 计算截止时间（毫秒时间戳）
            cutoff_timestamp = int((time.time() - request.days * 86400) * 1000)
            
            # 删除指定时间之前的记录
            cursor.execute("""
                DELETE FROM raw_data_records
                WHERE source_id = ? AND created_at < ?
            """, (source_id, cutoff_timestamp))
            
            deleted_count = cursor.rowcount
            
            # 更新数据源的总记录数
            cursor.execute("""
                SELECT COUNT(*) as count FROM raw_data_records
                WHERE source_id = ?
            """, (source_id,))
            
            remaining_count = cursor.fetchone()[0]
            
            cursor.execute("""
                UPDATE data_sources
                SET total_records = ?, updated_at = ?
                WHERE id = ?
            """, (remaining_count, int(time.time() * 1000), source_id))
            
            message = f"成功删除 {request.days} 天前的 {deleted_count} 条记录，剩余 {remaining_count} 条记录"
            
        elif request.clear_type == 'all':
            # 全部清空：删除所有记录并重置字段定义
            cursor.execute("""
                DELETE FROM raw_data_records
                WHERE source_id = ?
            """, (source_id,))
            
            deleted_count = cursor.rowcount
            
            # 重置数据源：清空字段定义、总记录数、同步次数等
            cursor.execute("""
                UPDATE data_sources
                SET field_schema = '{}',
                    total_records = 0,
                    sync_count = 0,
                    last_sync_time = NULL,
                    updated_at = ?
                WHERE id = ?
            """, (int(time.time() * 1000), source_id))
            
            message = f"成功清空所有数据（{deleted_count} 条记录）并重置表结构"
        
        else:
            conn.close()
            raise HTTPException(status_code=400, detail="无效的清理类型")
        
        conn.commit()
        conn.close()
        
        return {
            "code": 0,
            "message": message,
            "data": {
                "deleted_count": deleted_count,
                "clear_type": request.clear_type
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"清理失败：{str(e)}"}

# ========== 数据推送 API ==========

def verify_api_key(x_api_key: str = Header(...)):
    """验证API密钥"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, name, source_type, status FROM data_sources
        WHERE api_key = ? AND deleted = 0
    """, (x_api_key,))
    
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        raise HTTPException(status_code=401, detail="无效的API密钥")
    
    source = dict(row)
    
    if source['status'] != 'active':
        raise HTTPException(status_code=403, detail="数据源已停用")
    
    return source

@router.post("/push")
def push_data(
    request: DataPushRequest,
    source: Dict = Depends(verify_api_key)
):
    """推送数据到数据源"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        source_id = source['id']
        sync_time = datetime.now().isoformat()
        now = int(time.time() * 1000)
        
        # 创建同步日志
        log_id = f"log_{uuid.uuid4().hex[:12]}"
        start_time = sync_time
        
        success_count = 0
        failed_count = 0
        skipped_count = 0
        errors = []
        
        # 处理每条数据
        for idx, data_item in enumerate(request.data):
            try:
                # 生成记录ID
                record_id = f"rec_{uuid.uuid4().hex[:12]}"
                
                # 提取数据主键
                data_key = None
                if request.data_key_field and request.data_key_field in data_item:
                    data_key = str(data_item[request.data_key_field])
                
                # 如果是增量更新，检查是否已存在
                if request.incremental and data_key:
                    cursor.execute("""
                        SELECT id FROM raw_data_records
                        WHERE source_id = ? AND data_key = ? AND deleted = 0
                    """, (source_id, data_key))
                    
                    existing = cursor.fetchone()
                    
                    if existing:
                        # 更新现有记录
                        cursor.execute("""
                            UPDATE raw_data_records
                            SET raw_data = ?, sync_time = ?, updated_at = ?, version = version + 1
                            WHERE id = ?
                        """, (
                            json.dumps(data_item, ensure_ascii=False),
                            sync_time, now, existing[0]
                        ))
                        success_count += 1
                        continue
                
                # 插入新记录
                cursor.execute("""
                    INSERT INTO raw_data_records (
                        id, source_id, raw_data, data_key, data_type,
                        sync_time, is_processed, version, is_latest,
                        created_at, updated_at, deleted
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    record_id, source_id,
                    json.dumps(data_item, ensure_ascii=False),
                    data_key, source['source_type'],
                    sync_time, 0, 1, 1,
                    now, now, 0
                ))
                
                success_count += 1
                
            except Exception as e:
                failed_count += 1
                errors.append({
                    "index": idx,
                    "data": data_item,
                    "error": str(e)
                })
        
        # 更新数据源统计
        cursor.execute("""
            UPDATE data_sources
            SET total_records = total_records + ?,
                sync_count = sync_count + 1,
                last_sync_time = ?,
                updated_at = ?
            WHERE id = ?
        """, (success_count, sync_time, now, source_id))
        
        # 保存同步日志
        end_time = datetime.now().isoformat()
        status = 'success' if failed_count == 0 else ('partial' if success_count > 0 else 'failed')
        
        cursor.execute("""
            INSERT INTO data_sync_logs (
                id, source_id, sync_type, sync_method,
                total_count, success_count, failed_count, skipped_count,
                status, error_message, error_details,
                start_time, end_time, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            log_id, source_id, 'push', 'api',
            len(request.data), success_count, failed_count, skipped_count,
            status, f"成功{success_count}，失败{failed_count}" if failed_count > 0 else None,
            json.dumps(errors, ensure_ascii=False) if errors else None,
            start_time, end_time, now
        ))
        
        conn.commit()
        conn.close()
        
        return {
            "code": 0,
            "message": "数据推送完成",
            "data": {
                "total": len(request.data),
                "success": success_count,
                "failed": failed_count,
                "skipped": skipped_count,
                "log_id": log_id
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        return {"code": 1, "message": f"推送失败：{str(e)}"}

# ========== 数据查询 API ==========

@router.get("/{source_id}/records")
def get_records(
    source_id: str,
    page: int = 1,
    limit: int = 20,
    is_processed: Optional[bool] = None,
    search: Optional[str] = None
):
    """查询数据源的记录"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 构建查询
        query = """
            SELECT * FROM raw_data_records
            WHERE source_id = ? AND deleted = 0 AND is_latest = 1
        """
        params = [source_id]
        
        if is_processed is not None:
            query += " AND is_processed = ?"
            params.append(is_processed)
        
        if search:
            query += " AND (data_key LIKE ? OR raw_data LIKE ?)"
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern])
        
        # 获取总数
        count_query = query.replace("SELECT *", "SELECT COUNT(*)")
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        
        # 分页查询
        query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, (page - 1) * limit])
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # 转换结果
        records = []
        for row in rows:
            item = dict(row)
            # 解析 JSON 数据
            if item.get('raw_data'):
                item['raw_data'] = json.loads(item['raw_data'])
            records.append(item)
        
        conn.close()
        
        return {
            "code": 0,
            "data": records,
            "total": total,
            "page": page,
            "limit": limit
        }
        
    except Exception as e:
        return {"code": 1, "message": f"查询失败：{str(e)}"}

@router.get("/{source_id}/logs")
def get_sync_logs(source_id: str, page: int = 1, limit: int = 20):
    """获取同步日志"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取总数
        cursor.execute("""
            SELECT COUNT(*) FROM data_sync_logs WHERE source_id = ?
        """, (source_id,))
        total = cursor.fetchone()[0]
        
        # 分页查询
        cursor.execute("""
            SELECT * FROM data_sync_logs
            WHERE source_id = ?
            ORDER BY created_at DESC
            LIMIT ? OFFSET ?
        """, (source_id, limit, (page - 1) * limit))
        
        rows = cursor.fetchall()
        
        # 转换结果
        logs = []
        for row in rows:
            item = dict(row)
            if item.get('error_details'):
                item['error_details'] = json.loads(item['error_details'])
            logs.append(item)
        
        conn.close()
        
        return {
            "code": 0,
            "data": logs,
            "total": total,
            "page": page,
            "limit": limit
        }
        
    except Exception as e:
        return {"code": 1, "message": f"查询失败：{str(e)}"}

@router.get("/{source_id}/stats")
def get_stats(source_id: str, days: int = 7):
    """获取数据源统计"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取最近N天的统计
        cursor.execute("""
            SELECT * FROM data_source_stats
            WHERE source_id = ?
            ORDER BY stat_date DESC
            LIMIT ?
        """, (source_id, days))
        
        rows = cursor.fetchall()
        stats = [dict(row) for row in rows]
        
        conn.close()
        
        return {"code": 0, "data": stats}
        
    except Exception as e:
        return {"code": 1, "message": f"查询失败：{str(e)}"}


@router.post("/{source_id}/import-excel")
async def import_excel(
    source_id: str,
    file: UploadFile = File(...),
    incremental: bool = False
):
    """导入 Excel 文件"""
    try:
        # 检查 openpyxl 是否可用
        if not OPENPYXL_AVAILABLE:
            return {
                "code": 1,
                "message": "服务器未安装 openpyxl，无法解析 Excel 文件"
            }
        
        # 检查数据源是否存在
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM data_sources WHERE id = ?", (source_id,))
        source = cursor.fetchone()
        
        if not source:
            conn.close()
            return {"code": 1, "message": "数据源不存在"}
        
        source_dict = dict(source)
        
        # 读取 Excel 文件
        content = await file.read()
        
        # 使用 openpyxl 读取（不使用只读模式以支持更多属性）
        try:
            wb = load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            print(f"[Excel导入] 工作表名称: {ws.title}")
            print(f"[Excel导入] 最大行: {ws.max_row}, 最大列: {ws.max_column}")
        except Exception as e:
            conn.close()
            return {"code": 1, "message": f"Excel 文件读取失败：{str(e)}"}
        
        # 获取表头（第一行）
        rows = list(ws.rows)
        print(f"[Excel导入] 使用 ws.rows 读取到的行数: {len(rows)}")
        
        if len(rows) == 0:
            conn.close()
            return {"code": 1, "message": "Excel 文件为空"}
        
        # 打印第一行（表头）的原始数据
        print(f"[Excel导入] 第一行原始数据: {[cell.value for cell in rows[0]]}")
        
        headers = [cell.value for cell in rows[0] if cell.value]
        if not headers:
            conn.close()
            return {"code": 1, "message": "Excel 文件没有表头"}
        
        print(f"[Excel导入] 从 Excel 识别到的字段: {headers}")
        
        # 自动生成字段定义（根据 Excel 表头）
        auto_field_schema = {
            "fields": []
        }
        for header in headers:
            auto_field_schema["fields"].append({
                "name": header,
                "key": header,  # 使用表头作为 key
                "type": "text",  # 默认类型为文本
                "required": False
            })
        
        # 如果 Excel 的字段和原有字段定义不一致，自动更新数据源的字段定义
        print(f"[Excel导入] 自动更新数据源字段定义为: {auto_field_schema}")
        cursor.execute("""
            UPDATE data_sources
            SET field_schema = ?
            WHERE id = ?
        """, (json.dumps(auto_field_schema, ensure_ascii=False), source_id))
        
        # 转换数据
        records_data = []
        current_time = int(time.time() * 1000)
        
        print(f"[Excel导入] 表头: {headers}")
        print(f"[Excel导入] 总行数（含表头）: {len(rows)}")
        
        if len(rows) > 1:
            print(f"[Excel导入] 准备处理数据行，从第 2 行开始...")
        else:
            print(f"[Excel导入] ⚠️ 警告：只有表头行，没有数据行！")
        
        for row_idx, row in enumerate(rows[1:], start=2):  # 跳过表头
            # 打印原始行数据（前 10 个单元格）
            raw_row = [cell.value for cell in row]
            print(f"[Excel导入] 第 {row_idx} 行原始数据: {raw_row[:10]}")
            
            # 提取所有单元格值，不限制列数
            row_values = []
            has_data = False
            
            for cell in row:
                if cell.value is None or cell.value == '':
                    row_values.append(None)
                else:
                    value = str(cell.value).strip()
                    row_values.append(value)
                    if value and value != 'None':
                        has_data = True
            
            # 如果整行都没有数据，跳过
            if not has_data:
                print(f"[Excel导入] 跳过空行: 第 {row_idx} 行")
                continue
            
            print(f"[Excel导入] 处理第 {row_idx} 行，数据: {row_values[:5]}...")  # 只打印前5列
            
            # 构建数据对象 - 使用表头作为键
            data_obj = {}
            for idx, header in enumerate(headers):
                if idx < len(row_values):
                    data_obj[header] = row_values[idx]
                else:
                    data_obj[header] = None
            
            # 如果行的列数超过表头，也保存额外的列（使用列索引作为键）
            for idx in range(len(headers), len(row_values)):
                if row_values[idx] is not None:
                    data_obj[f"column_{idx+1}"] = row_values[idx]
            
            # 生成数据键（使用第一列作为键，如果第一列为空则使用行号）
            data_key = row_values[0] if (row_values and row_values[0]) else f"row_{row_idx}"
            
            record_id = str(uuid.uuid4())
            sync_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            records_data.append({
                'id': record_id,
                'source_id': source_id,
                'data_key': data_key,
                'raw_data': json.dumps(data_obj, ensure_ascii=False),
                'sync_time': sync_time,
                'is_processed': 0,
                'is_latest': 1,
                'created_at': current_time,
                'updated_at': current_time,
                'deleted': 0
            })
        
        wb.close()
        
        print(f"[Excel导入] 解析到有效记录数: {len(records_data)}")
        
        if not records_data:
            conn.close()
            return {"code": 1, "message": f"Excel 中没有有效数据（总行数: {len(rows)}, 表头: {headers}）"}
        
        # 如果是增量更新，标记旧数据
        if incremental:
            cursor.execute("""
                UPDATE raw_data_records
                SET is_latest = 0
                WHERE source_id = ? AND is_latest = 1
            """, (source_id,))
        else:
            # 全量覆盖，删除旧数据
            cursor.execute("""
                UPDATE raw_data_records
                SET deleted = 1
                WHERE source_id = ?
            """, (source_id,))
        
        # 插入新数据
        for record in records_data:
            cursor.execute("""
                INSERT INTO raw_data_records (
                    id, source_id, data_key, raw_data, sync_time,
                    is_processed, is_latest, created_at, updated_at, deleted
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                record['id'], record['source_id'], record['data_key'], 
                record['raw_data'], record['sync_time'], record['is_processed'], 
                record['is_latest'], record['created_at'], record['updated_at'], 
                record['deleted']
            ))
        
        # 更新数据源统计
        cursor.execute("""
            UPDATE data_sources
            SET total_records = (
                SELECT COUNT(*) FROM raw_data_records 
                WHERE source_id = ? AND deleted = 0 AND is_latest = 1
            ),
            sync_count = sync_count + 1,
            last_sync_time = ?,
            updated_at = ?
            WHERE id = ?
        """, (source_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), current_time, source_id))
        
        # 记录同步日志
        log_id = str(uuid.uuid4())
        sync_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            INSERT INTO data_sync_logs (
                id, source_id, sync_type, sync_method, 
                total_count, success_count, failed_count, skipped_count,
                status, start_time, end_time, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            log_id, source_id, 'manual', 'excel',
            len(records_data), len(records_data), 0, 0,
            'success', sync_time_str, sync_time_str, current_time
        ))
        
        conn.commit()
        conn.close()
        
        return {
            "code": 0,
            "message": f"成功导入 {len(records_data)} 条记录",
            "data": {
                "imported_count": len(records_data),
                "columns": headers
            }
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"code": 1, "message": f"导入失败：{str(e)}"}


# ========== 外部查询 API（供技术团队调用）==========

@router.get("/{source_id}/query")
def query_records(
    source_id: str,
    api_key: str = Header(..., alias="X-API-Key"),
    order_by: str = "updated_at",  # updated_at 或 order_no
    limit: int = 1,
    updated_after: Optional[str] = None  # 更新时间筛选（格式：2025-01-27 10:00:00）
):
    """
    外部查询接口（供技术团队调用）
    
    参数：
    - api_key: API密钥（Header: X-API-Key）
    - order_by: 排序字段（updated_at=按更新时间倒序, order_no=按订单号倒序）
    - limit: 返回记录数（默认1条）
    - updated_after: 筛选更新时间之后的记录（可选，格式：2025-01-27 10:00:00）
    
    返回：
    {
      "code": 0,
      "message": "查询成功",
      "data": [
        {
          "订单号": "...",
          "更新时间": "...",
          ...
        }
      ],
      "total": 100
    }
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 验证 API 密钥
        cursor.execute("""
            SELECT id, name FROM data_sources 
            WHERE id = ? AND api_key = ? AND status = 'active'
        """, (source_id, api_key))
        
        source = cursor.fetchone()
        if not source:
            conn.close()
            return {"code": 401, "message": "API密钥无效或数据源不存在"}
        
        # 构建查询
        query = """
            SELECT id, data_key, raw_data, sync_time, created_at, updated_at
            FROM raw_data_records
            WHERE source_id = ? AND deleted = 0 AND is_latest = 1
        """
        params = [source_id]
        
        # 按更新时间筛选
        if updated_after:
            try:
                # 转换为时间戳
                from datetime import datetime
                dt = datetime.strptime(updated_after, '%Y-%m-%d %H:%M:%S')
                timestamp = int(dt.timestamp() * 1000)
                query += " AND updated_at > ?"
                params.append(timestamp)
            except ValueError:
                conn.close()
                return {"code": 400, "message": "updated_after 格式错误，应为：2025-01-27 10:00:00"}
        
        # 排序
        if order_by == "order_no":
            # 按订单号倒序（从 raw_data 中提取）
            query += " ORDER BY json_extract(raw_data, '$.订单号') DESC"
        else:
            # 默认按更新时间倒序
            query += " ORDER BY updated_at DESC"
        
        # 限制返回数量
        query += f" LIMIT {min(limit, 100)}"  # 最多100条
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # 获取总数
        count_query = """
            SELECT COUNT(*) FROM raw_data_records
            WHERE source_id = ? AND deleted = 0 AND is_latest = 1
        """
        count_params = [source_id]
        
        if updated_after:
            count_query += " AND updated_at > ?"
            count_params.append(timestamp)
        
        cursor.execute(count_query, count_params)
        total = cursor.fetchone()[0]
        
        # 格式化返回数据
        result = []
        for record in records:
            record_dict = dict(record)
            
            # 解析 raw_data
            try:
                raw_data = json.loads(record_dict['raw_data'])
            except:
                raw_data = {}
            
            # 添加系统字段
            raw_data['_record_id'] = record_dict['id']
            raw_data['_sync_time'] = record_dict['sync_time']
            
            # 格式化时间戳
            from datetime import datetime
            raw_data['_created_at'] = datetime.fromtimestamp(record_dict['created_at'] / 1000).strftime('%Y-%m-%d %H:%M:%S')
            raw_data['_updated_at'] = datetime.fromtimestamp(record_dict['updated_at'] / 1000).strftime('%Y-%m-%d %H:%M:%S')
            
            result.append(raw_data)
        
        conn.close()
        
        return {
            "code": 0,
            "message": "查询成功",
            "data": result,
            "total": total,
            "count": len(result)
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"code": 1, "message": f"查询失败：{str(e)}"}
