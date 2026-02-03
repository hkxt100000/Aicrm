#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 字段分析工具
用于分析"测试智能表格.xlsx"的字段结构
"""

import openpyxl
import json
from datetime import datetime

def analyze_excel(file_path):
    """分析Excel文件的字段结构"""
    try:
        # 打开Excel文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        print("=" * 80)
        print("Excel 字段分析报告")
        print("=" * 80)
        print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"工作表名称: {sheet.title}")
        print(f"数据行数: {sheet.max_row}")
        print(f"数据列数: {sheet.max_column}")
        print("=" * 80)
        print()
        
        # 获取表头（第一行）
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value if cell_value else f"未命名列{col}")
        
        print(f"总字段数: {len(headers)}")
        print("=" * 80)
        print()
        
        # 字段详细信息
        print("字段列表:")
        print("-" * 80)
        
        field_details = []
        
        for idx, header in enumerate(headers, 1):
            # 获取该列的示例数据（前3行，跳过表头）
            sample_values = []
            for row in range(2, min(5, sheet.max_row + 1)):
                cell_value = sheet.cell(row=row, column=idx).value
                if cell_value is not None:
                    sample_values.append(str(cell_value))
            
            # 判断数据类型
            data_type = "未知"
            if sample_values:
                first_val = sample_values[0]
                if first_val.replace('.', '', 1).replace('-', '', 1).isdigit():
                    data_type = "数字"
                elif '/' in first_val or '-' in first_val:
                    try:
                        datetime.strptime(first_val, '%Y-%m-%d')
                        data_type = "日期"
                    except:
                        data_type = "文本"
                else:
                    data_type = "文本"
            
            field_info = {
                "序号": idx,
                "字段名": header,
                "推测类型": data_type,
                "示例值": sample_values[:3] if sample_values else ["无数据"]
            }
            
            field_details.append(field_info)
            
            # 打印字段信息
            print(f"{idx:3d}. {header}")
            print(f"     类型: {data_type}")
            print(f"     示例: {', '.join(sample_values[:3]) if sample_values else '无数据'}")
            print()
        
        print("=" * 80)
        print()
        
        # 生成JSON格式输出
        json_output = {
            "分析时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "工作表名称": sheet.title,
            "总字段数": len(headers),
            "数据行数": sheet.max_row - 1,  # 减去表头
            "字段列表": field_details
        }
        
        # 保存为JSON文件
        json_file = file_path.replace('.xlsx', '_字段分析.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_output, f, ensure_ascii=False, indent=2)
        
        print(f"✅ JSON报告已保存: {json_file}")
        print()
        
        # 生成文本格式输出（方便复制粘贴）
        txt_file = file_path.replace('.xlsx', '_字段分析.txt')
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Excel 字段分析报告\n")
            f.write("=" * 80 + "\n")
            f.write(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"工作表名称: {sheet.title}\n")
            f.write(f"总字段数: {len(headers)}\n")
            f.write(f"数据行数: {sheet.max_row - 1}\n")
            f.write("=" * 80 + "\n\n")
            
            f.write("字段列表（简洁版）:\n")
            f.write("-" * 80 + "\n")
            for field in field_details:
                f.write(f"{field['序号']:3d}. {field['字段名']} ({field['推测类型']})\n")
            
            f.write("\n" + "=" * 80 + "\n\n")
            
            f.write("字段详细信息:\n")
            f.write("-" * 80 + "\n")
            for field in field_details:
                f.write(f"\n{field['序号']:3d}. {field['字段名']}\n")
                f.write(f"     类型: {field['推测类型']}\n")
                f.write(f"     示例: {', '.join(field['示例值'])}\n")
        
        print(f"✅ 文本报告已保存: {txt_file}")
        print()
        
        # 生成SQL建表语句
        sql_file = file_path.replace('.xlsx', '_建表SQL.sql')
        with open(sql_file, 'w', encoding='utf-8') as f:
            f.write("-- 订单原始数据表\n")
            f.write("CREATE TABLE IF NOT EXISTS raw_orders (\n")
            f.write("    id TEXT PRIMARY KEY,\n")
            f.write("    source_id TEXT NOT NULL,  -- 数据源ID\n")
            f.write("    raw_data TEXT NOT NULL,   -- JSON格式完整原始数据\n")
            f.write("    \n")
            f.write("    -- 原始字段\n")
            
            for field in field_details:
                field_name = field['字段名'].replace(' ', '_').replace('/', '_').replace('-', '_')
                if field['推测类型'] == "数字":
                    sql_type = "REAL"
                elif field['推测类型'] == "日期":
                    sql_type = "TEXT"
                else:
                    sql_type = "TEXT"
                
                f.write(f"    {field_name} {sql_type},  -- {field['字段名']}\n")
            
            f.write("    \n")
            f.write("    -- 系统字段\n")
            f.write("    sync_time TEXT NOT NULL,      -- 同步时间\n")
            f.write("    is_processed BOOLEAN DEFAULT 0,  -- 是否已加工\n")
            f.write("    created_at INTEGER NOT NULL,\n")
            f.write("    updated_at INTEGER NOT NULL\n")
            f.write(");\n")
            f.write("\n")
            f.write("-- 创建索引\n")
            f.write("CREATE INDEX IF NOT EXISTS idx_raw_orders_source ON raw_orders(source_id);\n")
            f.write("CREATE INDEX IF NOT EXISTS idx_raw_orders_sync_time ON raw_orders(sync_time);\n")
            f.write("CREATE INDEX IF NOT EXISTS idx_raw_orders_processed ON raw_orders(is_processed);\n")
        
        print(f"✅ SQL建表语句已保存: {sql_file}")
        print()
        
        print("=" * 80)
        print("分析完成！")
        print("=" * 80)
        print()
        print("生成的文件：")
        print(f"  1. {json_file}  (JSON格式，包含完整信息)")
        print(f"  2. {txt_file}   (文本格式，方便阅读)")
        print(f"  3. {sql_file}   (SQL建表语句)")
        print()
        print("请把这3个文件的内容发给我！")
        
    except FileNotFoundError:
        print(f"❌ 错误：找不到文件 {file_path}")
        print()
        print("请确保文件路径正确，或者把Excel文件放在以下位置：")
        print(f"  {file_path}")
    except Exception as e:
        print(f"❌ 错误：{str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Excel文件路径（请根据实际情况修改）
    excel_file = "测试智能表格.xlsx"
    
    print()
    print("开始分析Excel文件...")
    print(f"文件路径: {excel_file}")
    print()
    
    analyze_excel(excel_file)
