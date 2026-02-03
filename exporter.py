"""
客户数据Excel导出模块
支持导出完整客户信息，包括头像下载、标签分类、跟进记录等
"""
import io
import json
import time
import requests
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import sqlite3

# 添加方式映射
ADD_WAY_MAP = {
    0: '未知',
    1: '扫描二维码',
    2: '搜索手机号',
    3: '名片分享',
    4: '群聊',
    5: '手机通讯录',
    6: '微信联系人',
    7: '添加好友申请',
    8: '第三方应用',
    9: '搜索邮箱',
    201: '内部成员共享',
    202: '管理员分配'
}

# 会话状态映射
IM_STATUS_MAP = {
    0: '正常',
    1: '已流失',
    2: '删除好友'
}

class CustomerExporter:
    """客户数据导出器"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        
    def export_to_excel(
        self,
        customer_ids: Optional[List[str]] = None,
        filters: Optional[Dict] = None,
        include_avatar: bool = True
    ) -> bytes:
        """
        导出客户数据到Excel
        
        Args:
            customer_ids: 要导出的客户ID列表（None表示导出所有）
            filters: 筛选条件
            include_avatar: 是否包含头像
            
        Returns:
            Excel文件的字节流
        """
        # 1. 查询客户数据
        customers = self._query_customers(customer_ids, filters)
        
        if not customers:
            raise ValueError("没有可导出的客户数据")
        
        # 2. 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "客户数据"
        
        # 3. 设置表头样式
        headers = self._get_headers()
        self._setup_headers(ws, headers)
        
        # 4. 填充数据
        for idx, customer in enumerate(customers, start=2):
            self._fill_customer_row(ws, idx, customer, include_avatar)
        
        # 5. 调整列宽
        self._adjust_column_widths(ws)
        
        # 6. 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _get_headers(self) -> List[str]:
        """获取表头列表"""
        return [
            '序号',
            '头像',
            '客户姓名',
            '性别',
            '客户类型',
            '所属员工',
            '添加时间',
            '添加方式',
            '会话状态',
            '手机号',
            '企业名称',
            '企业标签',
            '个人标签',
            '规则组标签',
            '描述信息',
            'UnionID',
            '渠道来源',
            '外部联系人ID',
            '最新跟进时间',
            '跟进次数'
        ]
    
    def _setup_headers(self, ws, headers: List[str]):
        """设置表头样式"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 冻结首行
        ws.freeze_panes = "A2"
    
    def _query_customers(
        self,
        customer_ids: Optional[List[str]],
        filters: Optional[Dict]
    ) -> List[Dict]:
        """查询客户数据"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # 构建查询SQL
        sql = """
            SELECT 
                c.*,
                COUNT(DISTINCT f.id) as follow_count,
                MAX(f.follow_time) as last_follow_time
            FROM customers c
            LEFT JOIN follow_records f ON c.id = f.customer_id
        """
        
        where_clauses = []
        params = []
        
        # ID筛选
        if customer_ids:
            placeholders = ','.join(['?' for _ in customer_ids])
            where_clauses.append(f"c.id IN ({placeholders})")
            params.extend(customer_ids)
        
        # 其他筛选条件
        if filters:
            if filters.get('owner_userid'):
                where_clauses.append("c.owner_userid = ?")
                params.append(filters['owner_userid'])
            
            if filters.get('search'):
                where_clauses.append("(c.name LIKE ? OR c.remark LIKE ?)")
                search_term = f"%{filters['search']}%"
                params.extend([search_term, search_term])
        
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        
        sql += " GROUP BY c.id ORDER BY c.add_time DESC"
        
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        
        customers = [dict(row) for row in rows]
        conn.close()
        
        return customers
    
    def _fill_customer_row(self, ws, row_idx: int, customer: Dict, include_avatar: bool):
        """填充客户数据行"""
        col_idx = 1
        
        # 序号
        ws.cell(row=row_idx, column=col_idx, value=row_idx - 1)
        col_idx += 1
        
        # 头像
        if include_avatar and customer.get('avatar'):
            try:
                avatar_img = self._download_and_resize_avatar(customer['avatar'])
                if avatar_img:
                    img = XLImage(avatar_img)
                    img.width = 40
                    img.height = 40
                    ws.add_image(img, f"B{row_idx}")
                    ws.row_dimensions[row_idx].height = 40
            except Exception as e:
                print(f"[导出] 下载头像失败: {e}")
        col_idx += 1
        
        # 客户姓名
        ws.cell(row=row_idx, column=col_idx, value=customer.get('name', ''))
        col_idx += 1
        
        # 性别
        gender_map = {0: '未知', 1: '男', 2: '女'}
        ws.cell(row=row_idx, column=col_idx, value=gender_map.get(customer.get('gender', 0), '未知'))
        col_idx += 1
        
        # 客户类型
        type_map = {1: '微信用户', 2: '企业微信用户'}
        ws.cell(row=row_idx, column=col_idx, value=type_map.get(customer.get('type', 1), '微信用户'))
        col_idx += 1
        
        # 所属员工
        ws.cell(row=row_idx, column=col_idx, value=customer.get('owner_name', ''))
        col_idx += 1
        
        # 添加时间
        add_time = customer.get('add_time', 0)
        if add_time:
            add_time_str = datetime.fromtimestamp(add_time).strftime('%Y-%m-%d %H:%M:%S')
        else:
            add_time_str = ''
        ws.cell(row=row_idx, column=col_idx, value=add_time_str)
        col_idx += 1
        
        # 添加方式
        add_way = customer.get('add_way', 0)
        ws.cell(row=row_idx, column=col_idx, value=ADD_WAY_MAP.get(add_way, '未知'))
        col_idx += 1
        
        # 会话状态
        im_status = customer.get('im_status', 0)
        ws.cell(row=row_idx, column=col_idx, value=IM_STATUS_MAP.get(im_status, '正常'))
        col_idx += 1
        
        # 手机号
        remark_mobiles = customer.get('remark_mobiles', '')
        if remark_mobiles:
            try:
                mobiles = json.loads(remark_mobiles)
                ws.cell(row=row_idx, column=col_idx, value=', '.join(mobiles))
            except:
                ws.cell(row=row_idx, column=col_idx, value=remark_mobiles)
        col_idx += 1
        
        # 企业名称
        ws.cell(row=row_idx, column=col_idx, value=customer.get('remark_corp_name', ''))
        col_idx += 1
        
        # 企业标签
        enterprise_tags = self._parse_tags(customer.get('enterprise_tags', ''))
        ws.cell(row=row_idx, column=col_idx, value=enterprise_tags)
        col_idx += 1
        
        # 个人标签
        personal_tags = self._parse_tags(customer.get('personal_tags', ''))
        ws.cell(row=row_idx, column=col_idx, value=personal_tags)
        col_idx += 1
        
        # 规则组标签
        rule_tags = self._parse_tags(customer.get('rule_tags', ''))
        ws.cell(row=row_idx, column=col_idx, value=rule_tags)
        col_idx += 1
        
        # 描述信息
        ws.cell(row=row_idx, column=col_idx, value=customer.get('description', ''))
        col_idx += 1
        
        # UnionID
        ws.cell(row=row_idx, column=col_idx, value=customer.get('unionid', ''))
        col_idx += 1
        
        # 渠道来源
        ws.cell(row=row_idx, column=col_idx, value=customer.get('state', ''))
        col_idx += 1
        
        # 外部联系人ID
        ws.cell(row=row_idx, column=col_idx, value=customer.get('id', ''))
        col_idx += 1
        
        # 最新跟进时间
        last_follow_time = customer.get('last_follow_time', 0)
        if last_follow_time:
            follow_time_str = datetime.fromtimestamp(last_follow_time).strftime('%Y-%m-%d %H:%M:%S')
        else:
            follow_time_str = '无'
        ws.cell(row=row_idx, column=col_idx, value=follow_time_str)
        col_idx += 1
        
        # 跟进次数
        ws.cell(row=row_idx, column=col_idx, value=customer.get('follow_count', 0))
    
    def _parse_tags(self, tags_json: str) -> str:
        """解析标签JSON为可读字符串"""
        if not tags_json:
            return ''
        try:
            tags = json.loads(tags_json)
            return ', '.join([tag.get('tag_name', '') for tag in tags if tag.get('tag_name')])
        except:
            return ''
    
    def _download_and_resize_avatar(self, avatar_url: str) -> Optional[io.BytesIO]:
        """下载并调整头像大小"""
        try:
            response = requests.get(avatar_url, timeout=10)
            if response.status_code == 200:
                img = Image.open(io.BytesIO(response.content))
                img = img.convert('RGB')
                img.thumbnail((80, 80), Image.LANCZOS)
                
                output = io.BytesIO()
                img.save(output, format='JPEG')
                output.seek(0)
                return output
        except Exception as e:
            print(f"[导出] 处理头像失败: {e}")
        return None
    
    def _adjust_column_widths(self, ws):
        """调整列宽"""
        column_widths = {
            'A': 6,   # 序号
            'B': 10,  # 头像
            'C': 15,  # 客户姓名
            'D': 8,   # 性别
            'E': 12,  # 客户类型
            'F': 15,  # 所属员工
            'G': 20,  # 添加时间
            'H': 15,  # 添加方式
            'I': 10,  # 会话状态
            'J': 15,  # 手机号
            'K': 20,  # 企业名称
            'L': 25,  # 企业标签
            'M': 25,  # 个人标签
            'N': 25,  # 规则组标签
            'O': 30,  # 描述信息
            'P': 30,  # UnionID
            'Q': 15,  # 渠道来源
            'R': 30,  # 外部联系人ID
            'S': 20,  # 最新跟进时间
            'T': 10,  # 跟进次数
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
